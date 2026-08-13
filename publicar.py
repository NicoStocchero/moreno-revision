#!/usr/bin/env python3
"""
Publica una revisión de cola de MORENO a partir del xlsx del barrido.

    python3 publicar.py barrido.xlsx --fecha 2026-08-14 --version 0.30.0 \
        --contenido contenido.json

Qué hace, en este orden:

1. Lee el xlsx y saca la hoja de leads.
2. Calcula solos los números del encabezado y del resumen.
3. Rellena plantilla.html.
4. Escribe /AAAA-MM-DD/index.html y copia esa misma página a la raíz.
5. Actualiza el bloque de archivo con todas las revisiones que existan.

No hace commit ni push: eso lo hace quien lo llame, para que quede en su historial.
"""
import json, sys, argparse, re, os, glob
from datetime import date

RAIZ = os.path.dirname(os.path.abspath(__file__))

CAMPOS = {
    'Cliente':'cli', 'Lead':'lead', 'Empresa':'emp', 'Sender y variante':'sender',
    'Nuestro último mensaje antes del suyo (literal)':'nuestro', 'Toque':'toque',
    'Último mensaje del lead (literal)':'lead_msg', 'Días desde ese mensaje':'dias',
    'Grupo':'g', 'Motivo del grupo':'gmot', 'Tag propuesto':'tag', 'Tag que tenía':'tag0',
    'Grupo de la matriz (normalizado)':'mat', 'Bloque de investigación':'inv',
    'Peldaño origen':'p0', 'Peldaño destino':'p1', 'Mensaje propuesto':'msg',
    'Preguntas que lleva':'preg', 'Frase de a qué nos dedicamos':'frase',
    'Tipo de cierre':'cierre', 'Estructura usada':'estr', 'Voz del sender aplicada':'voz',
    'Cadena de follow-up':'fu', 'Confianza':'conf', 'Camino de salida':'salida',
    'Defecto del mensaje real':'defecto', 'Veredicto contra el real':'ver',
    'Variante (normalizada)':'var',
}

def leer_xlsx(ruta):
    import openpyxl
    wb = openpyxl.load_workbook(ruta)
    hoja = None
    for n in wb.sheetnames:
        if 'lead' in n.lower(): hoja = wb[n]; break
    if hoja is None: hoja = wb[wb.sheetnames[0]]

    cab = [c.value for c in hoja[1]]
    # las dos columnas de fecha vienen con el mismo nombre; se distinguen por posición
    vistos, cols = {}, []
    for h in cab:
        if h in vistos:
            vistos[h] += 1; cols.append(f"{h}__{vistos[h]}")
        else:
            vistos[h] = 0; cols.append(h)

    filas = []
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if not fila or not fila[0]: continue
        cruda = {cols[i]: (fila[i] if i < len(fila) and fila[i] is not None else "")
                 for i in range(len(cols))}
        reg = {v: str(cruda.get(k, "")).strip() for k, v in CAMPOS.items()}
        fechas = [cruda.get(c, "") for c in cols if c.startswith('Fecha')]
        reg['fn'] = str(fechas[0]).strip() if len(fechas) > 0 else ""
        reg['fl'] = str(fechas[1]).strip() if len(fechas) > 1 else ""
        filas.append(reg)
    return filas

def numeros(filas):
    from collections import Counter
    g = Counter(f['g'] for f in filas)
    conf = Counter((f['conf'] or '').split(' ')[0] for f in filas if f['conf'])
    cli = Counter(f['cli'] for f in filas)
    return {
        'total': len(filas),
        'borradores': sum(1 for f in filas if f['msg']),
        'clientes': len(cli),
        'grupos': dict(g),
        'conf': dict(conf),
        'por_cliente': dict(cli),
    }

def bloque_controles(ctrl):
    """ctrl: lista de dicts con lab, val, note."""
    out = []
    for c in ctrl:
        clase = ' sm' if len(c['val']) > 9 else ''
        out.append(
            '<div class="ctl"><div class="lab">%s</div>'
            '<div class="val%s">%s</div><div class="note">%s</div></div>'
            % (c['lab'], clase, c['val'], c['note']))
    return "\n".join(out)

def bloque_cambios(cambios):
    """cambios: lista de dicts con titulo, estado, antes, ahora, cita (opcional)."""
    out = []
    for c in cambios:
        est = 't-hecho' if c.get('estado', 'Hecho').lower().startswith('hecho') else 't-pend'
        cita = ('<p class="quote">%s</p>' % c['cita']) if c.get('cita') else ''
        out.append(
            '<div class="cambio"><div class="tit"><h3>%s</h3>'
            '<span class="tag-mini %s">%s</span></div>'
            '<div class="antes"><h4>Antes</h4><p>%s</p></div>'
            '<div><h4>Ahora</h4><p>%s</p>%s</div></div>'
            % (c['titulo'], est, c.get('estado','Hecho'), c['antes'], c['ahora'], cita))
    return "\n".join(out)

def bloque_sugerencias(sugs):
    out = []
    for s in sugs:
        parrafos = "".join('<p>%s</p>' % p for p in s['texto'])
        out.append('<div class="sug"><h3>%s</h3>%s<p class="quien">%s</p></div>'
                   % (s['titulo'], parrafos, s['quien']))
    if not out: return ''
    return "\n".join(out)

def bloque_archivo(fecha_actual):
    fechas = sorted([os.path.basename(d) for d in glob.glob(os.path.join(RAIZ, '20*-*-*'))
                     if os.path.isdir(d)], reverse=True)
    if fecha_actual not in fechas: fechas.insert(0, fecha_actual)
    if len(fechas) < 2: return ''
    enlaces = " · ".join(
        ('<strong>%s</strong>' % f) if f == fecha_actual else '<a href="../%s/">%s</a>' % (f, f)
        for f in fechas)
    return ('<div class="legacy" style="margin-bottom:14px"><h3>Revisiones anteriores</h3>'
            '<p>%s</p></div>' % enlaces)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    ap.add_argument('--fecha', default=date.today().isoformat())
    ap.add_argument('--version', required=True, help='versión del motor, por ejemplo 0.30.0')
    ap.add_argument('--contenido', required=True,
                    help='JSON con controles, cambios y sugerencias de esta revisión')
    a = ap.parse_args()

    filas = leer_xlsx(a.xlsx)
    if not filas:
        print('ERROR: el xlsx no tiene filas de leads'); sys.exit(1)
    n = numeros(filas)
    cont = json.load(open(a.contenido, encoding='utf-8'))

    plantilla = open(os.path.join(RAIZ, 'plantilla.html'), encoding='utf-8').read()
    datos = json.dumps(filas, ensure_ascii=False, separators=(',', ':')).replace('</', '<\\/')

    dia, mes, anio = a.fecha[8:10], a.fecha[5:7], a.fecha[0:4]
    MESES = ['enero','febrero','marzo','abril','mayo','junio','julio','agosto',
             'septiembre','octubre','noviembre','diciembre']
    legible = '%s de %s de %s' % (int(dia), MESES[int(mes)-1], anio)

    sub = cont.get('sub') or (
        '%d hilos reales de %d clientes, %d respuestas redactadas. Cada uno con nuestro mensaje '
        'anterior delante. No se envió nada, no se etiquetó nada y no se tocó ningún CRM.'
        % (n['total'], n['clientes'], n['borradores']))

    html = (plantilla
        .replace('{{DATOS}}', datos)
        .replace('{{FECHA}}', legible)
        .replace('{{EYEBROW}}', 'Barrido en modo prueba · %s · sao v%s' % (legible, a.version))
        .replace('{{SUB}}', sub)
        .replace('{{CONTROLES}}', bloque_controles(cont['controles']))
        .replace('{{CAMBIOS}}', bloque_cambios(cont['cambios']))
        .replace('{{SUGERENCIAS}}', bloque_sugerencias(cont.get('sugerencias', [])))
        .replace('{{ARCHIVO}}', bloque_archivo(a.fecha)))

    quedan = re.findall(r'\{\{[A-Z]+\}\}', html)
    if quedan:
        print('ERROR: marcadores sin rellenar:', quedan); sys.exit(1)

    carpeta = os.path.join(RAIZ, a.fecha)
    os.makedirs(carpeta, exist_ok=True)
    open(os.path.join(carpeta, 'index.html'), 'w', encoding='utf-8').write(html)
    open(os.path.join(RAIZ, 'index.html'), 'w', encoding='utf-8').write(html)

    print('Publicada la revisión del %s' % legible)
    print('  hilos: %d · borradores: %d · clientes: %d' % (n['total'], n['borradores'], n['clientes']))
    print('  grupos: %s' % n['grupos'])
    print('  confianza: %s' % n['conf'])
    print('  ficheros: %s/index.html y index.html' % a.fecha)

if __name__ == '__main__':
    main()
